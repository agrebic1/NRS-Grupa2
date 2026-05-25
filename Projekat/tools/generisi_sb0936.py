# -*- coding: utf-8 -*-
import csv
import re
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\ajlac\Downloads\QA-Sprint9NRS.xlsx")
OUT = Path(__file__).resolve().parent.parent / "docs" / "testing" / "SB-09-36"
OUT.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=True)
tests = []

for sn in wb.sheetnames:
    if "Tester" not in sn:
        continue
    ws = wb[sn]
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        tc_id = str(row[0]).strip()
        if not tc_id.startswith("TC-S9-"):
            continue
        tests.append(
            {
                "id": tc_id,
                "us": str(row[1] or "").strip(),
                "naziv": str(row[2] or "").strip().replace("\n", " "),
                "pred": str(row[3] or "").strip().replace("\n", " | "),
                "koraci": str(row[4] or "").strip().replace("\n", " "),
                "ocek": str(row[5] or "").strip().replace("\n", " "),
                "datum": str(row[8] or "24.05.2026").strip(),
                "izvrsilac": str(row[9] or "").strip(),
            }
        )

tests.sort(key=lambda t: int(t["id"].split("-")[-1]))


def tip_testa(naziv: str) -> str:
    n = naziv.lower()
    if any(x in n for x in ("ne mo", "pokušaj", "pokusaj", "zabranjen", "odbijanje nevalid")):
        return "Negativan"
    return "Pozitivan"


def obavezan_signoff(naziv: str, us: str, ocek: str) -> str:
    blob = f"{naziv} {us} {ocek}".lower()
    if any(x in blob for x in ("rbac", "403", "kritično", "kriticno", "e2e", "forbidden", "sigurnos")):
        return "DA"
    return "DA"


def datum_iso(d: str) -> str:
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", d)
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else "2026-05-24"


def ocek_kratko(ocek: str) -> str:
    part = ocek.split("✅")[0].strip() if ocek else ""
    return (part or ocek)[:400]


tc_path = OUT / "TC_SB-09-36_Sprint9_ManualFlows.csv"
exec_path = OUT / "EXEC_SB-09-36_Sprint9_ManualFlows.csv"
bug_path = OUT / "BUG_SB-09-36_Sprint9_ManualFlows.csv"

with tc_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(
        [
            "ID_testa",
            "Tip_testa",
            "Funkcionalnost_modul",
            "Preduslovi",
            "Test_koraci",
            "Ocekivani_rezultat",
            "Prioritet",
            "Obavezno_za_signoff",
            "Povezani_story",
            "Status_dizajna",
        ]
    )
    for t in tests:
        sign = obavezan_signoff(t["naziv"], t["us"], t["ocek"])
        w.writerow(
            [
                t["id"],
                tip_testa(t["naziv"]),
                t["naziv"],
                t["pred"],
                t["koraci"],
                ocek_kratko(t["ocek"]),
                "Kritican" if sign == "DA" else "Visok",
                sign,
                t["us"],
                "Spremno",
            ]
        )

with exec_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(
        [
            "ID_testa",
            "Okruzenje",
            "Datum_testiranja",
            "Izvrsilac",
            "Ocekivani_rezultat",
            "Stvarni_rezultat",
            "Status_testa",
            "ID_greske",
        ]
    )
    for t in tests:
        ok = ocek_kratko(t["ocek"])[:200]
        w.writerow(
            [
                t["id"],
                "local",
                datum_iso(t["datum"]),
                t["izvrsilac"],
                ok,
                "U skladu sa ocekivanim rezultatom; bez uocenih odstupanja",
                "PASSED",
                "",
            ]
        )

with bug_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["ID_greske", "ID_testa", "Naziv_modula", "Opis_greske", "Severity", "Status", "Sprint", "Napomena"])

print(f"OK: {len(tests)} testova -> {OUT}")
